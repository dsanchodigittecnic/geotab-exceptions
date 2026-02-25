#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Geotab → Excel por GRUPO (pasado por parámetro) con detección automática de grupo (id / nombre)
y selección automática de RULE_IDS según el DIAGNOSTIC_ID.

Reglas:
- Si DIAGNOSTIC_ID == "DiagnosticPowerTakeoffEngagedId":
    RULE_IDS = [Recogida Basura, FMS, FMS group condition, TODO]
- Si DIAGNOSTIC_ID == "DiagnosticAux1Id":
    RULE_IDS = [Recogida Basura, AUX, AUX group condition, TODO]

Genera XLSX con:
- 00_Parametros
- 01_Resumen_grupo
- 02_Por_vehiculo
- 03_Segmentos_1
- 04_Excepciones

Uso:
  python geotab_excel.py b2B01
  python geotab_excel.py "Nombre del grupo"

Opciones:
  python geotab_excel.py b2B01 --diagnostic DiagnosticAux1Id
  python geotab_excel.py b2B01 --diagnostic DiagnosticPowerTakeoffEngagedId
  python geotab_excel.py b2B01 --from 2026-02-18T23:00:00.000Z --to 2026-02-19T22:59:59.000Z
  python geotab_excel.py b2B01 --include-subgroups
  python geotab_excel.py b2B01 --out salida.xlsx

Requisitos:
  pip install mygeotab python-dateutil openpyxl
"""

import sys
import os
import argparse
import datetime as pydt
from datetime import timezone
from dateutil import parser as dtparser
import mygeotab

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ==============================
# CONFIGURACIÓN (login)
# ==============================

SERVER = "my.geotab.com"
DATABASE = "emaya"
USERNAME = "dsancho@digittecnic.com"

# Recomendado: variable de entorno GEOTAB_PASSWORD
PASSWORD = os.getenv("GEOTAB_PASSWORD", "Catalunya4**")

DEFAULT_FROM_DATE = "2026-02-18T23:00:00.000Z"
DEFAULT_TO_DATE   = "2026-02-19T22:59:59.000Z"

# Diagnóstico por defecto

# b2B01
# DEFAULT_DIAGNOSTIC_ID = "DiagnosticPowerTakeoffEngagedId" #FMS

# b2B02
DEFAULT_DIAGNOSTIC_ID = "DiagnosticAux1Id" #AUX


# ==============================
# RULE SETS según DIAGNOSTIC
# ==============================

RULES_BY_DIAGNOSTIC = {
    "DiagnosticPowerTakeoffEngagedId": [
        "aMJv_sZ41UUSTPJqNHmcCAw",  # Recogida Basura
        "akWqirOGg_0mvKWrj3iACZQ",  # Test Barriendo/Recogida (FMS)
        "ayvMMtxMesEWI40RGFEYDqw",  # Test Barriendo/Recogida (FMS) - Group condition
        "aTMStgBucpUWDxX4ykxsblw"   # Test Barriendo/Recogida (TODO)
    ],
    "DiagnosticAux1Id": [
        "aMJv_sZ41UUSTPJqNHmcCAw",  # Recogida Basura
        "aCDCWlou1aUqfbbiwsGKdSg",  # Test Barriendo/Recogida (AUX)
        "az8GWV5Q7QEmL87I8A0Dmnw",  # Test Barriendo/Recogida (AUX) - Group condition
        "aTMStgBucpUWDxX4ykxsblw"   # Test Barriendo/Recogida (TODO)
    ]
}


# ==============================
# HELPERS
# ==============================

def parse_iso(x):
    if isinstance(x, pydt.datetime):
        if x.tzinfo is None:
            return x.replace(tzinfo=timezone.utc)
        return x.astimezone(timezone.utc)

    if isinstance(x, str):
        dt = dtparser.isoparse(x)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)

    raise TypeError("Fecha inválida")


def overlaps(a_start, a_end, b_start, b_end):
    return a_start < b_end and a_end > b_start


def is_one_value(x):
    if x in (1, 1.0, True):
        return True
    if isinstance(x, dict):
        if x.get("value") in (1, 1.0, True):
            return True
    return False


def seconds(dt_start, dt_end):
    return max(0.0, (dt_end - dt_start).total_seconds())


def connect():
    api = mygeotab.API(
        username=USERNAME,
        password=PASSWORD,
        database=DATABASE,
        server=SERVER
    )
    api.authenticate()
    return api


# ==============================
# GRUPOS / DISPOSITIVOS
# ==============================

def resolve_group_auto(api, group_input):
    """
    Auto-detect:
      1) si group_input es Group.id
      2) si no, buscar por Group.name exacto
    Devuelve: (group_id, group_name, how)
    """
    # 1) probar por id
    try:
        g = api.get("Group", search={"id": group_input})
        if g:
            return g[0]["id"], g[0].get("name", ""), "id"
    except Exception:
        pass

    # 2) probar por nombre exacto
    g = api.get("Group", search={"name": group_input})
    if g:
        return g[0]["id"], g[0].get("name", group_input), "name"

    raise ValueError("No se encontró el grupo (ni por id ni por nombre): %s" % group_input)


def get_devices_in_group(api, group_id, include_subgroups=False):
    """
    Devuelve lista de Device del grupo.
    include_subgroups: intento genérico (si tu cuenta expone Group.parent)
    """
    if not include_subgroups:
        return api.get("Device", search={"groups": [{"id": group_id}]})

    groups_to_use = [group_id]
    try:
        children = api.get("Group", search={"parent": {"id": group_id}})
        for g in children or []:
            groups_to_use.append(g["id"])
    except Exception:
        pass

    group_objs = [{"id": gid} for gid in groups_to_use]
    return api.get("Device", search={"groups": group_objs})


# ==============================
# REGLAS / EXCEPCIONES
# ==============================

def get_rule_name(api, rule_obj_or_id, cache):
    if isinstance(rule_obj_or_id, dict):
        if rule_obj_or_id.get("name"):
            return rule_obj_or_id["name"]
        rule_id = rule_obj_or_id.get("id")
    else:
        rule_id = rule_obj_or_id

    if rule_id in cache:
        return cache[rule_id]

    rules = api.get("Rule", search={"id": rule_id})
    if rules:
        name = rules[0]["name"]
        cache[rule_id] = name
        return name

    cache[rule_id] = rule_id
    return rule_id


def get_exceptions(api, device_id, from_date, to_date, rule_ids):
    events = []
    for rule_id in rule_ids:
        search = {
            "deviceSearch": {"id": device_id},
            "ruleSearch": {"id": rule_id},
            "fromDate": from_date,
            "toDate": to_date
        }
        data = api.get("ExceptionEvent", search=search)
        if data:
            events.extend(data)
    return events


def group_exceptions_by_rule(api, exceptions, rule_cache):
    counts = {}
    events_by_rule = {}
    for e in exceptions:
        name = get_rule_name(api, e["rule"], rule_cache)
        counts[name] = counts.get(name, 0) + 1
        events_by_rule.setdefault(name, []).append(e)
    return counts, events_by_rule


# ==============================
# MEDICIONES / SEGMENTOS
# ==============================

def get_measurements(api, device_id, diagnostic_id, from_date, to_date):
    search = {
        "deviceSearch": {"id": device_id},
        "diagnosticSearch": {"id": diagnostic_id},
        "fromDate": from_date,
        "toDate": to_date
    }
    data = api.get("StatusData", search=search)
    return sorted(data, key=lambda x: x["dateTime"])


def count_segments_value_1(status):
    intervals = []
    in_one = False
    start = None

    if not status or len(status) < 2:
        return 0, []

    for i in range(len(status) - 1):
        cur = status[i]
        nxt = status[i + 1]

        cur_dt = parse_iso(cur["dateTime"])
        nxt_dt = parse_iso(nxt["dateTime"])

        cur_is = is_one_value(cur.get("data"))
        nxt_is = is_one_value(nxt.get("data"))

        if cur_is and not in_one:
            start = cur_dt
            in_one = True

        if cur_is and not nxt_is and in_one:
            intervals.append((start, nxt_dt))
            in_one = False

    return len(intervals), intervals


def segments_overlap_by_rule(events_by_rule, intervals):
    result = {}
    total = len(intervals)

    for rule_name, events in events_by_rule.items():
        count = 0
        for seg_start, seg_end in intervals:
            for e in events:
                if overlaps(parse_iso(e["activeFrom"]), parse_iso(e["activeTo"]), seg_start, seg_end):
                    count += 1
                    break
        result[rule_name] = (total, count)

    return result


# ==============================
# EXCEL HELPERS
# ==============================

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20

def autosize_columns(ws, max_width=70):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

def add_table(ws, start_row, start_col, end_row, end_col, name):
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)


# ==============================
# CLI ARGS
# ==============================

def parse_args():
    ap = argparse.ArgumentParser(description="Geotab Excel por grupo (auto-detect id/nombre) + reglas según diagnóstico.")
    ap.add_argument("group", help="ID o NOMBRE del grupo (auto-detect).")
    ap.add_argument("--diagnostic", default=DEFAULT_DIAGNOSTIC_ID,
                    help="Diagnostic ID (ej: DiagnosticPowerTakeoffEngagedId o DiagnosticAux1Id)")
    ap.add_argument("--from", dest="from_date", default=DEFAULT_FROM_DATE,
                    help="fromDate ISO (UTC) ej: 2026-02-18T23:00:00.000Z")
    ap.add_argument("--to", dest="to_date", default=DEFAULT_TO_DATE,
                    help="toDate ISO (UTC) ej: 2026-02-19T22:59:59.000Z")
    ap.add_argument("--include-subgroups", action="store_true",
                    help="Intentar incluir subgrupos (depende de tu cuenta).")
    ap.add_argument("--out", dest="out_file", default="",
                    help="Nombre del XLSX de salida (opcional).")
    return ap.parse_args()


# ==============================
# MAIN
# ==============================

def main():
    args = parse_args()

    if args.diagnostic not in RULES_BY_DIAGNOSTIC:
        valid = ", ".join(RULES_BY_DIAGNOSTIC.keys())
        print("\nERROR: diagnostic no soportado:", args.diagnostic)
        print("Diagnósticos soportados:", valid)
        return 2

    diagnostic_id = args.diagnostic
    rule_ids = RULES_BY_DIAGNOSTIC[diagnostic_id]

    api = connect()

    group_id, group_name, how = resolve_group_auto(api, args.group)
    devices = get_devices_in_group(api, group_id, include_subgroups=args.include_subgroups)

    if not devices:
        print("No se encontraron vehículos en el grupo:", args.group, "| resuelto:", group_name, "| id:", group_id)
        return 1

    rule_cache = {}

    # Nombres de reglas (para que siempre aparezcan aunque 0)
    rule_names = [get_rule_name(api, rid, rule_cache) for rid in rule_ids]

    rows_por_vehiculo = []
    rows_segmentos = []
    rows_excepciones = []

    total_segments_group = 0
    total_exceptions_by_rule_group = {rn: 0 for rn in rule_names}
    total_overlap_by_rule_group = {rn: 0 for rn in rule_names}

    print("Grupo input:", args.group, "| Detectado como:", how)
    print("Grupo:", group_name, "| id:", group_id)
    print("Diagnostic:", diagnostic_id)
    print("Vehículos:", len(devices))
    print("Rango:", args.from_date, "->", args.to_date)

    for d in devices:
        device_id = d["id"]
        device_name = d.get("name", device_id)

        exceptions = get_exceptions(api, device_id, args.from_date, args.to_date, rule_ids)
        measurements = get_measurements(api, device_id, diagnostic_id, args.from_date, args.to_date)

        segments_count, intervals = count_segments_value_1(measurements)
        counts, events_by_rule = group_exceptions_by_rule(api, exceptions, rule_cache)
        segment_stats = segments_overlap_by_rule(events_by_rule, intervals)

        total_segments_group += segments_count

        # Detalle segmentos
        for (s, e) in intervals:
            dur_s = seconds(s, e)
            rows_segmentos.append([
                device_name, device_id,
                s.replace(tzinfo=None), e.replace(tzinfo=None),
                dur_s, dur_s / 60.0
            ])

        # Detalle excepciones
        for ex in exceptions:
            rn = get_rule_name(api, ex["rule"], rule_cache)
            s = parse_iso(ex["activeFrom"])
            e = parse_iso(ex["activeTo"])
            dur_s = seconds(s, e)
            rows_excepciones.append([
                device_name, device_id, rn,
                s.replace(tzinfo=None), e.replace(tzinfo=None),
                dur_s, dur_s / 60.0
            ])

        # Por vehículo x regla (siempre todas las del diagnóstico seleccionado)
        for rn in rule_names:
            exc_count = int(counts.get(rn, 0))
            overlap_count = int(segment_stats.get(rn, (segments_count, 0))[1]) if segments_count else 0
            pct = (overlap_count / segments_count) if segments_count else 0.0

            rows_por_vehiculo.append([
                device_name, device_id, rn,
                segments_count, exc_count, overlap_count, pct
            ])

            total_exceptions_by_rule_group[rn] += exc_count
            total_overlap_by_rule_group[rn] += overlap_count

        print("Procesado:", device_name, "| segmentos:", segments_count, "| excepciones:", len(exceptions))

    # ========= crear Excel =========
    wb = Workbook()

    # 00_Parametros
    ws0 = wb.active
    ws0.title = "00_Parametros"
    ws0.append(["Campo", "Valor"])
    ws0.append(["Grupo (input)", args.group])
    ws0.append(["Detectado como", how])
    ws0.append(["Grupo (name)", group_name])
    ws0.append(["Grupo (id)", group_id])
    ws0.append(["Include subgrupos", str(args.include_subgroups)])
    ws0.append(["Desde (UTC)", args.from_date])
    ws0.append(["Hasta (UTC)", args.to_date])
    ws0.append(["Diagnostic ID", diagnostic_id])
    ws0.append(["Reglas (nombres)", ", ".join(rule_names)])
    style_header_row(ws0, 1)
    ws0.freeze_panes = "A2"
    autosize_columns(ws0)

    # 01_Resumen_grupo
    ws1 = wb.create_sheet("01_Resumen_grupo")
    ws1.append(["Regla", "Excepciones (grupo)", "Segmentos valor=1 (grupo)", "Segmentos con solape", "% solape (solape/segmentos)"])
    style_header_row(ws1, 1)

    for rn in rule_names:
        exc = total_exceptions_by_rule_group.get(rn, 0)
        ov = total_overlap_by_rule_group.get(rn, 0)
        ws1.append([rn, exc, total_segments_group, ov, None])

    for r in range(2, 2 + len(rule_names)):
        ws1.cell(row=r, column=5).value = f"=IF(C{r}=0,0,D{r}/C{r})"
        ws1.cell(row=r, column=5).number_format = "0.0%"

    ws1.freeze_panes = "A2"
    autosize_columns(ws1)
    add_table(ws1, 1, 1, ws1.max_row, ws1.max_column, "TablaResumenGrupo")

    # 02_Por_vehiculo
    ws2 = wb.create_sheet("02_Por_vehiculo")
    ws2.append(["Vehículo", "Device ID", "Regla", "Segmentos valor=1", "Excepciones", "Segmentos con solape", "% solape"])
    style_header_row(ws2, 1)

    for row in rows_por_vehiculo:
        ws2.append(row)

    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=7).number_format = "0.0%"

    ws2.freeze_panes = "A2"
    autosize_columns(ws2)
    add_table(ws2, 1, 1, ws2.max_row, ws2.max_column, "TablaPorVehiculo")

    # 03_Segmentos_1
    ws3 = wb.create_sheet("03_Segmentos_1")
    ws3.append(["Vehículo", "Device ID", "Inicio segmento", "Fin segmento", "Duración (s)", "Duración (min)"])
    style_header_row(ws3, 1)

    for row in rows_segmentos:
        ws3.append(row)

    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=3).number_format = "yyyy-mm-dd hh:mm:ss"
        ws3.cell(row=r, column=4).number_format = "yyyy-mm-dd hh:mm:ss"
        ws3.cell(row=r, column=6).number_format = "0.00"

    ws3.freeze_panes = "A2"
    autosize_columns(ws3)
    if ws3.max_row >= 2:
        add_table(ws3, 1, 1, ws3.max_row, ws3.max_column, "TablaSegmentos")

    # 04_Excepciones
    ws4 = wb.create_sheet("04_Excepciones")
    ws4.append(["Vehículo", "Device ID", "Regla", "Inicio excepción", "Fin excepción", "Duración (s)", "Duración (min)"])
    style_header_row(ws4, 1)

    for row in rows_excepciones:
        ws4.append(row)

    for r in range(2, ws4.max_row + 1):
        ws4.cell(row=r, column=4).number_format = "yyyy-mm-dd hh:mm:ss"
        ws4.cell(row=r, column=5).number_format = "yyyy-mm-dd hh:mm:ss"
        ws4.cell(row=r, column=7).number_format = "0.00"

    ws4.freeze_panes = "A2"
    autosize_columns(ws4)
    if ws4.max_row >= 2:
        add_table(ws4, 1, 1, ws4.max_row, ws4.max_column, "TablaExcepciones")

    # Guardar
    safe_group = "".join([c if c.isalnum() else "_" for c in (group_name or args.group)])
    out_name = args.out_file.strip()
    if not out_name:
        out_name = f"geotab_{safe_group}_{args.from_date[:10]}_{args.to_date[:10]}_{diagnostic_id}.xlsx"

    wb.save(out_name)
    print("\nExcel generado:", out_name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
